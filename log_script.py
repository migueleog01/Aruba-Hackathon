import pandas
from itertools import islice, groupby
import openpyxl

#function that takes in a string and a dictionary and returns a list of matched phrases
def search_keywords(input_string, dict):
    matched_phrases = []
    #print(dict)
    for keyword,phrases in dict.items():
        #for phrase in phrases:
            #print(phrase.strip)
            if keyword in input_string:
                #we dont want duplicates
                if keyword not in matched_phrases:
                    matched_phrases.append(keyword)
    return matched_phrases


#function that puts commands into a list of lists
def iap_or_capFunction(type_command):
    dataframe = openpyxl.load_workbook("Log_Database.xlsx")
    worksheet = dataframe.active
    #CHANGE TO 1 FOR HACKATHON
    index  = 1
    primarylist = []
    commandlist = []
    prev_index = None
    if type_command == "IAP":
        for row in islice(worksheet.iter_rows(), 1, None):
    #if the row has a value and isnt None
            if row[2].value is not None:
                stringt = row[2].value
                #the format of the string is "0 show ap debug radio-stats X | include MCS"
                current_index = int(stringt[0])
                #we set prev index to our current
                if prev_index is None:
                    prev_index = current_index
                #if the current is the same as the prev meaning if its 0 = 0, we append it to a list with commands
                if current_index == prev_index:
                    commandlist.append(stringt)
                else: #else meaning its not the same, we are done with that commandlist so we push it to the primarylist, thatll be a list of lists
                    primarylist.append(commandlist)
                    commandlist = [stringt]
                    prev_index = current_index #and we set previndex to the current

        # Append the last commandlist if there are any remaining commands
        if commandlist:
            primarylist.append(commandlist)
        
        
        #return 0
    elif type_command == "CAP":
        #iterate through our rows       
        for row in islice(worksheet.iter_rows(), 1, None):
            #if the row has a value and isnt None
            if row[3].value is not None:
                stringt = row[3].value
                #the format of the string is "0 show ap debug radio-stats X | include MCS"
                current_index = int(stringt[0])
                #we set prev index to our current
                if prev_index is None:
                    prev_index = current_index
                #if the current is the same as the prev meaning if its 0 = 0, we append it to a list with commands
                if current_index == prev_index:
                    commandlist.append(stringt)
                else: #else meaning its not the same, we are done with that commandlist so we push it to the primarylist, thatll be a list of lists
                    primarylist.append(commandlist)
                    commandlist = [stringt]
                    prev_index = current_index #and we set previndex to the current

        # Append the last commandlist if there are any remaining commands
        if commandlist:
            primarylist.append(commandlist)
        
        
    return primarylist


def excelWords_list(worksheet):
    list_of_lists = []
    colindex = 1
    #for row in worksheet.iter_rows():
    index = 1
    for row in islice(worksheet.iter_rows(), 1, None):
        temp = ""
        
        if row[1].value is not None:

            database_string = row[1].value
            #split by the string and add it to list
            split_items = database_string.split(',')
            list_of_lists.append(split_items)
            #print(row[2].value)
            #temp = row[2].value
            #print(temp)
        #in our excel sheet there are values where row[1] is None such as A3 and A5
        if row[2].value is not None:
            #print(row[2].value)
            temp = row[2].value
            #print(temp)
    
    return list_of_lists
    

def createDictionary(list_of_lists, primarylist):
    mydict = {}

    listcount = 0
    for list in list_of_lists:
        counter = 0
        
        for item in list:
            
            #print(counter)
            #rint(item,listcount)

            #print(item, primarylist[counter])
            mydict[item] = primarylist[listcount]
            counter +=1
        listcount +=1
    return mydict



def printCommandsSteps(mydict, listofKeywords ):
    for word in listofKeywords:
        counter = 1
        if word in mydict:
            #print(len(mydict[word]) )
            #print( "The keyword is: ", word , "The commands are " , mydict[word])
            
            commandstring = mydict[word]
            print("Use these commands: ",word)
            print("")
            for command in commandstring:
                #replace the first character in command. replace it with counter variable
                command = command.replace(command[0], str(counter))
                #print set variable to just the step without first character
                command = command[2:]


                print(command)
                counter += 1
        print("")

def inputFunction():
    logsummary = input("Enter error Log Summary: ")
    iap_cap  = input("Enter IAP or CAP: ")
    while iap_cap != "IAP" and iap_cap != "CAP":
        iap_cap  = input("Enter IAP or CAP: ")
    else:
        exit
    
    return logsummary, iap_cap


#open the excel sheet
dataframe = openpyxl.load_workbook("Log_Database.xlsx")
worksheet = dataframe.active


#create a list of lists of commands
inputList = inputFunction()
input_string = inputList[0]
iap_cap = inputList[1]
primarylist = iap_or_capFunction(iap_cap)



#create a list of lists of keywords
list_of_lists = excelWords_list(worksheet)
#print(list_of_lists)

#create a dictionary with the keywords as the key and the list of commands as the value
mydict = createDictionary(list_of_lists, primarylist)
print(mydict)


print("")


#we get string from user and try to match it with our dictionary
listofKeywords = search_keywords(input_string, mydict)


#we print the commands and steps for the user that match the keywords from the string
printCommandsSteps(mydict, listofKeywords )





        
        

